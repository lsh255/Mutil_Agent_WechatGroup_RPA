from openpyxl import load_workbook
from typing import Dict, Any
from pathlib import Path


class UpdateExcelTool:
    """Excel更新工具：用于更新台账数据"""
    
    def __init__(self, template_path: str):
        """初始化Excel工具
        
        Args:
            template_path: Excel模板文件路径
        """
        self.template_path = Path(template_path)
        self.workbook = None
    
    def _load_workbook(self):
        """加载Excel工作簿"""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Excel模板文件不存在: {self.template_path}")
        
        self.workbook = load_workbook(self.template_path)
    
    def _find_or_create_row(self, sheet, user: str, task_id: str) -> int:
        """查找或创建数据行
        
        Args:
            sheet: 工作表对象
            user: 用户名
            task_id: 任务ID
            
        Returns:
            行号
        """
        # 假设第一列是用户名，第二列是任务ID
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value == user and row[1].value == task_id:
                return row_idx
        
        # 未找到，创建新行
        new_row_idx = sheet.max_row + 1
        return new_row_idx
    
    def update(self, data: Dict[str, Any]) -> bool:
        """更新Excel数据
        
        Args:
            data: 包含更新数据的字典，应包含:
                - user: 用户名
                - task_id: 任务ID
                - phase: 任务阶段
                - completed: 是否完成
                - summary: 内容摘要
                - extracted_info: 提取的信息
                
        Returns:
            是否更新成功
        """
        try:
            self._load_workbook()
            
            # 获取或创建工作表
            sheet = self.workbook.active
            
            # 查找或创建行
            row_idx = self._find_or_create_row(sheet, data["user"], data["task_id"])
            
            # 更新数据
            sheet.cell(row=row_idx, column=1, value=data["user"])
            sheet.cell(row=row_idx, column=2, value=data["task_id"])
            sheet.cell(row=row_idx, column=3, value=data["phase"])
            sheet.cell(row=row_idx, column=4, value=data["completed"])
            sheet.cell(row=row_idx, column=5, value=data["summary"])
            
            # 更新提取的信息（假设从第6列开始）
            for idx, (key, value) in enumerate(data["extracted_info"].items(), start=6):
                sheet.cell(row=row_idx, column=idx, value=value)
            
            # 保存工作簿
            self.workbook.save(self.template_path)
            return True
            
        except Exception as e:
            print(f"Excel更新失败: {e}")
            return False
        finally:
            if self.workbook:
                self.workbook.close()
    
    def get_all_data(self) -> list[Dict[str, Any]]:
        """获取Excel中所有数据
        
        Returns:
            数据列表
        """
        try:
            self._load_workbook()
            sheet = self.workbook.active
            
            data = []
            for row in sheet.iter_rows(min_row=2):
                if row[0].value:  # 跳过空行
                    data.append({
                        "user": row[0].value,
                        "task_id": row[1].value,
                        "phase": row[2].value,
                        "completed": row[3].value,
                        "summary": row[4].value,
                    })
            
            return data
            
        except Exception as e:
            print(f"获取Excel数据失败: {e}")
            return []
        finally:
            if self.workbook:
                self.workbook.close()
